import os,pickle,pprint,sys,csv,json
import youtube_dl
import spotipy
import tidalapi
from spotipy.oauth2 import SpotifyOAuth
from dotenv import load_dotenv
from youtube_api import YoutubeAPI
import openpyxl

# Funciones auxiliares
from common_functions import create_download_folder

class TidalAPI:
    def __init__(self):
        # Carga de las variables de entorno desde el archivo .env
        load_dotenv()
        self.session_file = "tidal_credentials.pickle"
        # Credenciales de las variables de entorno para utiliar API de Spotify
        self.TIDAL_CLIENT_ID = os.getenv('TIDAL_CLIENT_ID')
        self.TIDAL_CLIENT_SECRET = os.getenv('TIDAL_CLIENT_SECRET')
        self.TIDAL_REDIRECT_URI = os.getenv('TIDAL_REDIRECT_URI')

    def service_login(self):
        try:
            if os.path.exists(self.session_file):
                # Si el archivo de sesión existe, cargar la sesión desde el archivo
                with open(self.session_file, 'rb') as f:
                    session = pickle.load(f)
            else:
                # Si el archivo de sesión no existe, iniciar una nueva sesión y guardarla en el archivo
                session = tidalapi.Session()
                session.login_oauth_simple()
                with open(self.session_file, 'wb') as f:
                    pickle.dump(session, f)
        except Exception as e:
            print(f"Error loading or saving the session: {e}")
            session = None
        return session

    # Listar las playlists del usuario en Tidal
    def list_playlists(self):
        counter = 0
        tracks = 0
        session = self.service_login()

        playlists = session.user.playlists()
        print(f"PLAYLISTS:{playlists}")

        print()
        print(f'My Tidal playlists:\n')
        for playlist in playlists:
            counter+=1
            tracks += playlist.num_tracks
            print(f'* {playlist.name} (id: {playlist.id})')
        
        print()
        print(f'Playlists found: {counter}')
        print(f'Total tracks found:{tracks}')

    # Lista los tracks con su información de una playlist en particular de Tidal
    def get_playlist_data(self, playlist_name):
        session = self.service_login()
        playlists = session.user.playlists()
        playlist_data = []
        for playlist in playlists:
            if playlist.name == playlist_name:
                for track in playlist.tracks():
                    playlist_data.append({
                    'track_name': track.name,
                    'artist': track.artist.name,
                    'album_name': track.album.name,
                    'album_release_date': track.album.year,
                    'playlist':playlist_name,
                    'service':'Tidal'
                })
        return playlist_data


    # Exporta la información de una playlist de Tidal en formáto CSV
    def get_playlist_csv(self,playlist_name):
        data_csv = [['track_title','artist','album','year','playlist','service']]
        session = self.service_login()
        playlist_data = self.get_playlist_data(playlist_name)
        for track in playlist_data:
            track_title = track['track_name']
            artist = track['artist']
            album = track['album_name']
            year = track['album_release_date']
            data_csv.append([track_title,artist,album,year,playlist_name,'Tidal'])

        file_path = os.path.join(create_download_folder(self,'csv'), playlist_name + "_tidal.csv")

        try:
            with open(file_path, mode='w', newline='', encoding='utf-8') as file_csv:
                escritor_csv = csv.writer(file_csv)
                for row in data_csv:
                    escritor_csv.writerow(row)

            print(f"CSV file '{file_path}' successfully created")
        except Exception as e:
            print(f"Error creating the CSV file: {e}")
    
    # Exporta la información de una playlist de Tidal en formáto JSON
    def get_playlist_json(self,playlist_name):
        data_json = []
        session = self.service_login()
        playlist_data = self.get_playlist_data(playlist_name)
        for track in playlist_data:
            track_title = track['track_name']
            artist = track['artist']
            album = track['album_name']
            year = track['album_release_date']
            data_json.append({'track_title':track_title,'artist':artist,'album':album,'year':year,'playlist':playlist_name,'service':'Tidal'})

        file_path = os.path.join(create_download_folder(self,'json'), playlist_name + "_tidal.json")
        try:
            with open(file_path, "w", encoding="utf-8") as file:
                json.dump(data_json, file, indent=4, ensure_ascii=False)
                print(f"JSON file '{file_path}' successfully created")
        except Exception as e:
            print(f"Error creating the JSON file: {e}")
        
    # Exporta la información de una playlist de Tidal en formáto XLSX
    def get_playlist_xlsx(self,playlist_name):
        data_xlsx = []
        session = self.service_login()
        playlist_data = self.get_playlist_data(playlist_name)
        for track in playlist_data:
            track_title = track['track_name']
            artist = track['artist']
            album = track['album_name']
            year = track['album_release_date']
            data_xlsx.append({'track_title':track_title,'artist':artist,'album':album,'year':year,'playlist':playlist_name,'service':'Tidal'})
        # Crear un nuevo libro de trabajo de Excel
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = playlist_name + ' tidal playlist info'

            # Escribir encabezados
            sheet.append(["Track title", "Artist", "Album", "Year","Playlist","Service"])

            # Escribir datos de la playlist
            for track in data_xlsx:
                sheet.append([track["track_title"], track["artist"], track["album"],track["year"],track["playlist"],track["service"]])

            # Guardar el libro de trabajo en un archivo Excel
            file = os.path.join(create_download_folder(self,'xlsx'), playlist_name + "_tidal.xlsx")
            wb.save(file)
            print(f"XLSX file successfully created")
        except Exception as e:
            print(f"Error creating the Excel file: {e}")

    # Importa la información de una playlist de Tidal desde un archivo local en formáto JSON
    def upload_playlist_from_json(self,playlist_name):
        local_path = os.path.dirname(__file__)
        sub_directories = ['downloads', 'json']  # Lista de subdirectorios
        folder_path = os.path.join(local_path, *sub_directories)
        json_file = os.path.join(folder_path,playlist_name + '_tidal.json')
        try:
            with open(json_file, 'r', encoding='utf-8') as file:
                # Carga el contenido del archivo JSON en un diccionario
                playlist_data = json.load(file)

                session = self.service_login()

                playlist_description = "A playlist with songs uploaded with playMigrate"
                playlist = session.user.create_playlist(playlist_name, playlist_description)

                track_ids = []

                # Agrega las canciones a la playlist
                for track_info in playlist_data:
                    print(f'{track_info["track_title"]} {track_info["artist"]}')
                    # Busca la pista por nombre y artista
                    query = f'{track_info["track_title"]} {track_info["artist"]}'
                    tracks = session.search(query)
                    # Agrego el id de la pista para agregar mas tarde a la playlist
                    track_ids.append(tracks['tracks'][0].id)
            playlist.add(track_ids)
            print('Tidal playlist created successfully!')
        
        except FileNotFoundError:
            print("The file was not found")
        except json.decoder.JSONDecodeError:
            print("The file does not have a valid JSON format")
        except Exception as e:
            print("An error occurred while trying to read the JSON file:", e)
        
    # Importa la información de una playlist de Tidal desde un archivo local en formáto CSV
    def upload_playlist_from_csv(self,playlist_name):
        local_path = os.path.dirname(__file__)
        sub_directories = ['downloads', 'csv']  # Lista de subdirectorios
        folder_path = os.path.join(local_path, *sub_directories)
        csv_file = os.path.join(folder_path, playlist_name + '_tidal.csv')
    
        try:
            with open(csv_file, 'r', encoding='utf-8') as file:
                # Lee el archivo CSV
                csv_reader = csv.DictReader(file)
                session = self.service_login()
                playlist_description = "A playlist with songs uploaded with playMigrate"
                playlist = session.user.create_playlist(playlist_name, playlist_description)

                track_ids = []

                # Lee cada fila del archivo CSV
                for row in csv_reader:
                    track_title = row['track_title']
                    artist = row['artist']
                    print(f'{track_title} {artist}')
                    # Busca la pista por nombre y artista
                    query = f'{track_title} {artist}'
                    tracks = session.search(query)
                    # Agrega el id de la pista para agregar más tarde a la playlist
                    track_ids.append(tracks['tracks'][0].id)
                
                playlist.add(track_ids)
                print('Tidal playlist created successfully!')
        
        except FileNotFoundError:
            print("The file was not found")
        except Exception as e:
            print("An error occurred while trying to read the CSV file:", e)

    # Importa la información de una playlist de Tidal desde un archivo local en formáto XLSX
    def upload_playlist_from_xlsx(self,playlist_name):
        local_path = os.path.dirname(__file__)
        sub_directories = ['downloads', 'xlsx']  # Lista de subdirectorios
        folder_path = os.path.join(local_path, *sub_directories)
        xlsx_file = os.path.join(folder_path, playlist_name + '_tidal.xlsx')
    
        try:
            # Abre el archivo xlsx
            workbook = openpyxl.load_workbook(xlsx_file)
            sheet = workbook.active

            session = self.service_login()
            playlist_description = "A playlist with songs uploaded with playMigrate"
            playlist = session.user.create_playlist(playlist_name, playlist_description)
            track_ids = []

            # Itera sobre cada fila en la hoja de cálculo
            for row in sheet.iter_rows(values_only=True):
                track_title, artist = row[0], row[1]
                print(f'{track_title} {artist}')
                # Busca la pista por nombre y artista
                query = f'{track_title} {artist}'
                tracks = session.search(query)
            
                # Agrega el id de la pista para agregar más tarde a la playlist
                track_ids.append(tracks['tracks'][0].id)
            
            playlist.add(track_ids)
            print('Tidal playlist created successfully!')
        
        except FileNotFoundError:
            print("The file was not found")
        except Exception as e:
            print("An error occurred while trying to read the XLSX file:", e)

    # Retorna información del usuario de Tidal
    def user_info(self):
        session = self.service_login()
        user_info = session.user
        print((' ' + user_info.first_name + ' Tidal account information: ').center(100,'*'))
        
        print("Username:", user_info.first_name)
        print("Surname:", user_info.last_name)
        print("User ID:", user_info.id)
        print("Email:", user_info.email)
        
        print(('Favorites tracks: ').center(100,'*'))
        tracks = user_info.favorites.tracks()
        for track in tracks:
            print('* ' + track.name, "-", track.artist.name)

        print(('Favorites albums: ').center(100,'*'))
        albums = user_info.favorites.albums()
        for album in albums:
            print('* ' + album.name, "-", album.artist.name)

    # Creando una playlist de Tidal desde cero
    def create_playlist(self, playlist_name):
        try:
            session = self.service_login()
            playlist_description = "A playlist created with playMigrate"
            playlist = session.user.create_playlist(playlist_name, playlist_description)
            print(f'Playlist created successfully with ID: {playlist.id}')
            return playlist
        except Exception as e:
            print(f'Error while trying to create a playlist: {str(e)}')
    
    # Lista los tracks de una playlist en particular de Tidal
    def search_playlist_tracks(self, playlist_name):
        session = self.service_login()
        playlist_data = self.get_playlist_data(playlist_name)
        print()
        print(f'Playlist \"{playlist_name}\" with {len(playlist_data)} tracks found: \n')
        for track in playlist_data:
            track_title = track['track_name']
            artist = track['artist']
            album = track['album_name']
            year = track['album_release_date']
            print(f'* {track_title} by {artist} from the album  {album} released on {year}')

    # Migración de una playlist de Spotify a Tidal
    def migrate_playlist_from_sp(self, spotify_api, playlist_name):
        sp_playlist_data = spotify_api.get_playlist_data(playlist_name)
        session = self.service_login()
        # Creamos la playlist vacia
        playlist = self.create_playlist(playlist_name)
        track_ids = []
        # Agrega las canciones a la playlist
        for track_info in sp_playlist_data['tracks']:
            track_title = track_info['track_name']
            artist = track_info['artist']
            # Busca la pista por nombre y artista
            query = f'{track_title} {artist}'
            try:
                tracks = session.search(query)
                if tracks and 'tracks' in tracks and tracks['tracks']:
                    # Agrego el id de la pista para agregar mas tarde a la playlist
                    track_ids.append(tracks['tracks'][0].id)
                else:
                    print(f"The track '{track_title}' of '{artist}' was not found in Tidal")
            except Exception as e:
                print(f"Error occurred while searching for '{track_title}' of '{artist}' in Tidal: {e}")
        # Añadir las pistas encontradas a la playlist en Tidal
        try:
            playlist.add(track_ids)
            print('Tidal playlist created successfully!')
        except Exception as e:
            print(f"Error occurred while adding tracks to Tidal playlist: {e}")

    # Migración de una playlist de Youtube Music a Tidal
    def migrate_playlist_from_yt(self, youtube_api, playlist_name):
        print('Starting the playlist migration process...')
        session = self.service_login()
        # Creamos la playlist vacia
        playlist = self.create_playlist(playlist_name)
        track_ids = []
        tracks_to_migrate = []
        added_tracks = []
        # Avisamos al usuario que la playlist que se quiere migrar no existe en Youtube Music
        if youtube_api.get_playlist_id_by_name(playlist_name) is None:
            print(f'The entered playlist "{playlist_name}" does not exist')
            sys.exit()

        # Obtener la info de las pistas de la lista de reproducción de YouTube Music
        tracks_info = youtube_api.get_playlist_tracks(playlist_name)
        # Agrega las canciones a la playlist
        for track in tracks_info:
            track_id = track['track_id']
            try:
                with youtube_dl.YoutubeDL({}) as ydl:
                    video_info = ydl.extract_info(f"https://www.youtube.com/watch?v={track['track_id']}", download=False)
                # Extraer el nombre del artista y el título del álbum de la información del video
                artist = video_info.get('artist')
                track_title = video_info.get('title')
                year = video_info.get('release_year')  # Obtener el año de lanzamiento
                genre = video_info.get('genre')  # Obtener el género

            except youtube_dl.utils.DownloadError as e:
                print(f'Error processing the video {track["track_id"]}')
                print('Continuing with the next video...')
                print()
                continue

            except Exception as e:
                print(f'An unexpected error occurred')
                print('Continuing with the next video...')
                print()
                continue

            # Busca la pista por nombre y artista
            query = f'{track_title} {artist} {year} {genre}'

            try:
                tracks = session.search(query)
                if tracks and 'tracks' in tracks and tracks['tracks']:
                    # Agrego el id de la pista para agregar mas tarde a la playlist
                    track_ids.append(tracks['tracks'][0].id)
                else:
                    print(f"The track '{track_title}' of '{artist}' was not found in Tidal")
            except Exception as e:
                print(f"Error occurred while searching for '{track_title}' of '{artist}' in Tidal")
        # Añadir las pistas encontradas a la playlist en Tidal
        try:
            playlist.add(track_ids)
            print('Tidal playlist created successfully!')
        except Exception as e:
            print(f"Error occurred while adding tracks to Tidal playlist: {e}")



